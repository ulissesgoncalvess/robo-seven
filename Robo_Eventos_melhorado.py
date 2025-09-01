import ttkbootstrap as tb
from ttkbootstrap.constants import *
import subprocess
import threading
import os
from PIL import Image, ImageTk  # Para manipular imagem
from tkinter import BOTH
import sys

def executar_funcao():
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from openpyxl import Workbook, load_workbook
    from datetime import date, timedelta, datetime
    from tkinter.filedialog import asksaveasfilename
    import time
    import os
    import re
    import sys
    import tkinter as tk
    from tkinter import simpledialog

    # --- CONFIGURAÇÃO ---
    EXCEL_PATH = asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="Salvar planilha como"
    )
    USER = "emanuele@sevensuprimentos.com.br"
    PASS = "*Eas251080"

    root = tk.Tk()
    root.withdraw()
    data_usuario = simpledialog.askstring(
        title="Input",
        prompt="Digite a data desejada no formato DDMMAA (ex: 190825 para 19/08/25):"
    )
    if data_usuario and len(data_usuario.strip()) == 6 and data_usuario.isdigit():
        HOJE = f"{data_usuario[:2]}/{data_usuario[2:4]}/{data_usuario[4:]}"
    else:
        raise ValueError("Data inválida! Use o formato DDMMAA, ex: 190825")
    root.destroy()

    ESTADOS = [
        'AC', 'AL', 'AP', 'AM', 'BA', 'CE', 'DF', 'ES', 'GO', 'MA', 'MT', 'MS', 'MG',
        'PA', 'PB', 'PR', 'PE', 'PI', 'RJ', 'RN', 'RS', 'RO', 'RR', 'SC', 'SP', 'SE', 'TO'
    ]

    # --- PREPARA PLANILHA ---
    if os.path.exists(EXCEL_PATH):
        os.remove(EXCEL_PATH)
    wb = Workbook()
    ws = wb.active
    ws.title = "Eventos"
    ws.append(["Numero do evento", "UF(VALE)", "DATA", "DESCRIÇÃO", "QTDE", "UNID. MED", "pagina de descrição"])
    wb.save(EXCEL_PATH)

    # --- INICIA SELENIUM ---
    driver = webdriver.Chrome()
    wait = WebDriverWait(driver, 10)
    driver.get("https://vale.coupahost.com/sessions/supplier_login")

    # login
    wait.until(EC.presence_of_element_located((By.ID, "user_login")))
    driver.find_element(By.ID, "user_login").send_keys(USER)
    driver.find_element(By.ID, "user_password").send_keys(PASS, Keys.RETURN)

    # Clica no elemento de data duas vezes
    try:
        time_filter = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ch_start_time"]')))
        time_filter.click()
        time.sleep(5)
        time_filter = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ch_start_time"]')))
        time_filter.click()
    except:
        pass

    # Robo irá buscar todos os casos que a data inicio = data atual, até a ENCONTRAR ONTEM
    count = 1
    countBotao = 4
    encontrou_ontem = False
    while True:
        time.sleep(5)
        tbody = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="quote_request_table_tag"]')))
        print("tabela encontrada")
        linhas = tbody.find_elements(By.TAG_NAME, "tr")
        print(f"✅ Encontradas {len(linhas)} linhas na tabela.")

        # --- Buscar os números do evento ---
        for linha in linhas:
            try:
                colunas = linha.find_elements(By.TAG_NAME, "td")
                if not colunas or len(colunas) < 7:
                    continue

                # pula a linha se existir o ícone amarelo (flag_yellow) em qualquer lugar da linha
                yellow_flags = linha.find_elements(By.CSS_SELECTOR, "img[src*='flag_yellow']")
                if yellow_flags:
                    print("Pulando linha porque contém flag_yellow")
                    continue

                data_inicio = colunas[2].text.strip()

                if data_inicio < HOJE:
                    encontrou_ontem = True
                    print(f"❌ Encontrou data anterior a HOJE ({HOJE}): {data_inicio}. Parando a coleta.")
                    break

                if data_inicio != HOJE:
                    print(f"⚠️ Data {data_inicio} não é igual a HOJE ({HOJE}). Ignorando linha.")
                    continue
                numero_evento = colunas[0].find_element(By.TAG_NAME, "a").text.strip()
                data_final = colunas[3].text.strip()
                print(f"Número do evento: {numero_evento} | Data final: {data_final}")

                ws.append([numero_evento, '', data_final, '', '', '', ''])

            except Exception as e:
                print(f"⚠️ Não foi possível extrair dados da linha: {e}")

        if encontrou_ontem:
            break

        try:
            proximo = driver.find_element(By.CLASS_NAME, "next_page")
            print("✅ Botão 'Avançar' encontrado, clicando...")
        except:
            print("⚠️ Botão 'Avançar' não encontrado.")

        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", proximo)
        time.sleep(1)

        if encontrou_ontem:
            break

        try:
            proximo.click()
            print("✅ Botão 'Avançar' clicado.")
            time.sleep(3)
        except Exception as e:
            print(f"Não tem mais páginas ou erro ao clicar no botão 'Avançar': {e}")
            break

    wb.save(EXCEL_PATH)
    print(f"💾 Planilha salva em: {EXCEL_PATH}")

    # --- DETALHA CADA EVENTO ---
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Eventos"]

    for row in ws.iter_rows(min_row=2):
        evento = row[0].value
        driver.get(f"https://vale.coupahost.com/quotes/external_responses/{evento}/edit")
        wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))

        # --- VERIFICA EXISTÊNCIA DA PÁGINA DE DESCRIÇÃO ---
        try:
            botoes1 = driver.find_elements(By.XPATH, '//*[@id="pageContentWrapper"]/div[3]/div[2]/a[2]/span')
            if not botoes1:
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                botoes2 = driver.find_elements(By.ID, 'quote_response_submit')
                botoes2[0].click()
        except:
            row[6].value = "Erro ao verificar página de descrição"

        # Scroll e abre seção das informações
        try:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            wait.until(EC.presence_of_element_located((By.CLASS_NAME, "s-expandLines")))
            elementos = driver.find_elements(By.CLASS_NAME, "s-expandLines")
            elementos[0].click()
        except:
            pass

        # quantidade
        try:
            quantidade = driver.find_element(By.XPATH, '//*[@id="itemsAndServicesApp"]/div/div/div[1]/div[2]/div[2]/div/form/div/div/div[2]/div/div[2]/div/p/span[1]').text
            row[4].value = quantidade
        except:
            row[4].value = 'Não foi possivel coletar a quantidade'

        # unidade
        try:
            unidade = driver.find_element(By.XPATH, '//*[@id="itemsAndServicesApp"]/div/div/div[1]/div[2]/div[2]/div/form/div/div/div[2]/div/div[2]/div/p/span[2]').text
            row[5].value = unidade
        except:
            row[5].value = 'Não foi possivel coletar a unidade'

        # descrição
        try:
            descri = driver.find_element(By.XPATH, f'//*[@id="itemsAndServicesApp"]/div/div/div[1]/div[2]/div[2]/div/form/div/div/div[1]/div/div[2]/div/p').text
            desejado = re.search(r'PT\s*\|\|\s*(.*?)\*{3,}', descri, re.DOTALL)
            if desejado:
                row[3].value = desejado.group(1).strip()
            else:
                row[3].value = descri
        except:
            pass

        # UF
        try:
            uf_text = driver.find_element(By.XPATH, f'//*[@id="itemsAndServicesApp"]/div/div/div[1]/div[2]/div[2]/div/form/div/div/div[1]/div/div[8]/div/ul/li[1]/span').text
            for sig in ESTADOS:
                if sig in uf_text:
                    row[1].value = sig
                    break
        except:
            row[1].value = 'Não foi possivel coletar a UF'

        wb.save(EXCEL_PATH)

    driver.quit()
    print("Concluído! Planilha em:", EXCEL_PATH)



# --- INTERFACE ---
janela = tb.Window(themename="flatly")
janela.title("Robô de Eventos - Seven")
janela.geometry("800x400")
janela.resizable

frame = tb.Frame(janela, padding=20)
frame.pack(fill=BOTH, expand=True)

# --- Adiciona a LOGO ---
LOGO_PATH = r"C:\Users\Pessoal\Desktop\meu projeto\logo.png"
try:
    imagem = Image.open(LOGO_PATH)
    imagem = imagem.resize((200, 65))
    imagem_tk = ImageTk.PhotoImage(imagem)

    label_imagem = tb.Label(frame, image=imagem_tk)
    label_imagem.image = imagem_tk
    label_imagem.pack(pady=(0, 10))
except Exception as e:
    print(f"Erro ao carregar imagem: {e}")

# Título
titulo = tb.Label(frame, text="Robô de Eventos Seven", font=("Segoe UI", 18, "bold"))
titulo.pack(pady=(0, 20))

# Botão de iniciar
botao_iniciar = tb.Button(frame, text="Iniciar Robô", bootstyle=SUCCESS, width=30, command=executar_funcao)
botao_iniciar.pack(pady=5)

# Status
status_var = tb.StringVar(value="Aguardando início...")
status_label = tb.Label(frame, textvariable=status_var, bootstyle=INFO)
status_label.pack(pady=(20, 0))

janela.mainloop()
